"""Формулобезопасная запись УМР в xlsx: один лист формата тестового датасета."""

import math
from decimal import Decimal
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from ..publish import PublishedUmr


def _missing(value: object) -> bool:
    return value is None or (isinstance(value, float) and math.isnan(value))


def _scalar(value: object) -> object | None:
    if _missing(value):
        return None
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, "item") and not isinstance(value, str):
        return value.item()
    return value


def _write_frame(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title)
    for column, name in enumerate(frame.columns, start=1):
        worksheet.cell(1, column, str(name))
    for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column, value in enumerate(row, start=1):
            value = _scalar(value)
            cell = worksheet.cell(row_index, column)
            cell.value = value
            if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
                cell.data_type = "s"


def export_umr_workbook(umr: PublishedUmr, path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_frame(workbook, umr.sheet_name, umr.frame)
    workbook.save(path)
