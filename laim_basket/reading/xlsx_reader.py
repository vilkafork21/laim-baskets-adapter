"""Чтение xlsx в RawSheet: значения-объекты + карта формул + merged ranges.

Принципиально НЕ используем pd.read_excel: pandas-коэрция типов губит id
(float scientific notation), прячет формулы и merged cells. Грид читается
openpyxl дважды: data_only=True даёт кэшированные значения формул, второй
проход собирает карту формул (по ней доказываются футеры-агрегаты и volatile
RAND()-id).
"""

from dataclasses import dataclass, field

import openpyxl

from ..errors import ReadError


@dataclass
class RawSheet:
    """Лист «как есть»: прямоугольный грид без интерпретации."""

    name: str
    grid: list[list]                 # [row][col], 0-based; кэш значений формул
    merged: list[tuple[int, int, int, int]]  # (r1, c1, r2, c2) 0-based включительно
    formulas: dict[tuple[int, int], str] = field(default_factory=dict)

    @property
    def n_rows(self) -> int:
        return len(self.grid)

    @property
    def n_cols(self) -> int:
        return len(self.grid[0]) if self.grid else 0


def _used_bounds(rows: list[tuple], merged, formulas) -> tuple[int, int]:
    """Реальные границы данных: max индекс строки/колонки со значением,
    формулой или merged-диапазоном. Спасает от «раздутых» max_row/max_col
    (форматирование пустых ячеек, хвостовые пустые строки)."""
    last_row, last_col = -1, -1
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            if value is not None and value != "":
                last_row = max(last_row, r)
                last_col = max(last_col, c)
    for r1, c1, r2, c2 in merged:
        last_row = max(last_row, r2)
        last_col = max(last_col, c2)
    for r, c in formulas:
        last_row = max(last_row, r)
        last_col = max(last_col, c)
    return last_row, last_col


def _sheet_formulas(ws) -> dict[tuple[int, int], str]:
    """Карта формул листа: (row, col) 0-based → текст формулы."""
    return {
        (cell.row - 1, cell.column - 1): str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.data_type == "f"
    }


def read_workbook(path) -> dict[str, RawSheet]:
    """Прочитать все листы книги в RawSheet."""
    try:
        wb_values = openpyxl.load_workbook(path, data_only=True)
        wb_formulas = openpyxl.load_workbook(path, data_only=False)
    except Exception as exc:  # openpyxl кидает разнотипные ошибки на битых файлах
        raise ReadError(f"Не удалось открыть xlsx: {path}: {exc}", path=str(path)) from exc

    sheets: dict[str, RawSheet] = {}
    for ws in wb_values.worksheets:
        rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        merged = [
            (rng.min_row - 1, rng.min_col - 1, rng.max_row - 1, rng.max_col - 1)
            for rng in ws.merged_cells.ranges
        ]
        formulas = _sheet_formulas(wb_formulas[ws.title])
        last_row, last_col = _used_bounds(rows, merged, formulas)
        grid = [
            [rows[r][c] if c < len(rows[r]) else None for c in range(last_col + 1)]
            for r in range(last_row + 1)
        ]
        sheets[ws.title] = RawSheet(
            name=ws.title,
            grid=grid,
            merged=merged,
            formulas=formulas,
        )

    if not sheets:
        raise ReadError(f"В книге нет листов: {path}", path=str(path))
    return sheets
