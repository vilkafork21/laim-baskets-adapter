"""Превращение проверенной шапки и границ в прямоугольный регион таблицы."""

from dataclasses import dataclass, field

from openpyxl.utils import get_column_letter

from ..errors import ReadError
from ..reading.headers import dedup_names, header_merge_map, merge_header
from ..reading.xlsx_reader import RawSheet


@dataclass
class TableRegion:
    columns: list[str]
    rows: list[list[object]]
    source_rows: list[int]
    first_data_row0: int
    raw_indexes: list[int]
    column_letters: list[str]
    dedup_renames: dict[str, str] = field(default_factory=dict)
    accounting: dict[str, object] = field(default_factory=dict)

    def raw_index(self, column: str) -> int:
        return self.raw_indexes[self.columns.index(column)]


def build_region(sheet: RawSheet, header_rows: list[int], last_data_row: int) -> TableRegion:
    header_rows0 = [row - 1 for row in header_rows]
    first_data0 = max(header_rows0) + 1
    last_data0 = last_data_row - 1
    if first_data0 > last_data0 or last_data0 >= sheet.n_rows:
        raise ReadError(
            "Некорректные доказанные границы таблицы",
            first_data_row=first_data0 + 1,
            last_data_row=last_data_row,
            sheet_rows=sheet.n_rows,
        )
    merge_map = header_merge_map(sheet, header_rows0)
    raw_names = [merge_header(sheet, header_rows0, column, merge_map) for column in range(sheet.n_cols)]
    names = [name or f"unnamed_{get_column_letter(column + 1)}" for column, name in enumerate(raw_names)]
    names, renames = dedup_names(names)
    rows = [list(sheet.grid[row]) for row in range(first_data0, last_data0 + 1)]
    return TableRegion(
        columns=names,
        rows=rows,
        source_rows=list(range(first_data0 + 1, last_data0 + 2)),
        first_data_row0=first_data0,
        raw_indexes=list(range(sheet.n_cols)),
        column_letters=[get_column_letter(column + 1) for column in range(sheet.n_cols)],
        dedup_renames=renames,
        accounting={
            "sheet_rows": sheet.n_rows,
            "header_rows": len(header_rows),
            "first_data_row": first_data0 + 1,
            "last_data_row": last_data_row,
            "data_rows": len(rows),
        },
    )
