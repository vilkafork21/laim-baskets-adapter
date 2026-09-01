"""Щедрый физический снимок книги для LLM-разметки: факты, не решения.

Размер снимка ограничен константами defaults и не растёт с числом строк
корзины: модель видит якорные строки и статистику колонок, а не данные.
"""
from __future__ import annotations

import re
from collections import Counter

from openpyxl.utils import get_column_letter

from .. import defaults
from ..reading.xlsx_reader import RawSheet

_BLOB_MARKER = re.compile(defaults.BLOB_MARKER_PATTERN)


def _text(value: object) -> str:
    return "" if value is None else " ".join(str(value).split())


def _cap(value: object, limit: int) -> str:
    text = _text(value)
    return text if len(text) <= limit else text[: limit - 1] + "…"


def _is_numeric(value: object) -> bool:
    if isinstance(value, bool):
        return False
    try:
        float(_text(value).replace(" ", "").replace(",", ".").rstrip("%"))
        return bool(_text(value))
    except ValueError:
        return False


def _header_candidate(sheet: RawSheet) -> int:
    """Первая строка с двумя и более непустыми ячейками (0-based)."""
    for index, row in enumerate(sheet.grid):
        if sum(bool(_text(value)) for value in row) >= 2:
            return index
    return 0


def _anchor_rows(sheet: RawSheet) -> list[dict[str, object]]:
    """Первые непустые строки + строки смены профиля заполненности."""
    anchors: list[dict[str, object]] = []
    previous: tuple[bool, ...] | None = None
    for index, row in enumerate(sheet.grid):
        profile = tuple(bool(_text(value)) for value in row)
        if not any(profile):
            previous = profile
            continue
        is_opening = len(anchors) < defaults.EVIDENCE_ANCHOR_ROWS
        if (is_opening or profile != previous) and \
                len(anchors) < 2 * defaults.EVIDENCE_ANCHOR_ROWS:
            anchors.append({
                "row": index + 1,
                "cells": [_cap(value, defaults.EVIDENCE_CELL_CAP) for value in row],
            })
        previous = profile
    return anchors


def _column_snapshots(sheet: RawSheet, header: int) -> list[dict[str, object]]:
    columns = []
    rows = sheet.grid[header + 1:]
    for column in range(sheet.n_cols):
        present = [row[column] for row in rows if _text(row[column])]
        unique = Counter(_text(value) for value in present)
        item: dict[str, object] = {
            "address": get_column_letter(column + 1),
            "header": _text(sheet.grid[header][column])
            or f"unnamed_{get_column_letter(column + 1)}",
            "non_null": len(present),
            "numeric": sum(_is_numeric(value) for value in present),
            "samples": [
                _cap(value, defaults.EVIDENCE_CELL_CAP)
                for value in present[: defaults.EVIDENCE_SAMPLES_PER_COLUMN]
            ],
            "formulas": sum(1 for (_, col) in sheet.formulas if col == column),
        }
        if 0 < len(unique) <= defaults.EVIDENCE_UNIQUES_MAX:
            item["uniques"] = dict(unique.most_common())
        columns.append(item)
    return columns


def _dialogue_markers(columns: list[dict[str, object]]) -> list[str]:
    """Подсказка (не решение): маркеры реплик, физически найденные в сэмплах.

    Точное написание критично (АГЕНТ и АГЕНТЫ — разные маркеры), поэтому
    модель получает найденные строки дословно, а не факт их наличия."""
    found: dict[str, None] = {}
    for column in columns:
        markers = Counter(
            marker
            for value in column["samples"]
            for marker in _BLOB_MARKER.findall(str(value))
        )
        if len(markers) >= 2:
            for marker in markers:
                found.setdefault(marker)
    return list(found)


def _merged_ranges(sheet: RawSheet) -> list[str]:
    return [
        f"{get_column_letter(c1 + 1)}{r1 + 1}:{get_column_letter(c2 + 1)}{r2 + 1}"
        for r1, c1, r2, c2 in sheet.merged
    ]


def workbook_evidence(sheets: dict[str, RawSheet]) -> dict[str, object]:
    snapshots = []
    for sheet in sheets.values():
        header = _header_candidate(sheet)
        columns = _column_snapshots(sheet, header)
        snapshots.append({
            "name": sheet.name,
            "rows": sheet.n_rows,
            "columns_count": sheet.n_cols,
            "header_candidate_row": header + 1,
            "anchor_rows": _anchor_rows(sheet),
            "columns": columns,
            "merged": _merged_ranges(sheet),
            "dialogue_markers": _dialogue_markers(columns),
        })
    return {"sheets": snapshots}
