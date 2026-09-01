"""Ограниченный физический снимок книги для вывода layout — без решений."""

import re
from collections import Counter

from openpyxl.utils import get_column_letter

from .. import defaults
from ..reading.headers import header_merge_map, merge_header
from ..reading.xlsx_reader import RawSheet

_BLOB_MARKER = re.compile(defaults.BLOB_MARKER_PATTERN)


def _text(value: object) -> str:
    return "" if value is None else " ".join(str(value).split())


def _cap(value: object, limit: int) -> str:
    text = _text(value)
    return text if len(text) <= limit else text[: limit - 1] + "…"


def _header_candidate(sheet: RawSheet) -> int:
    for index, row in enumerate(sheet.grid):
        if sum(bool(_text(value)) for value in row) >= 2:
            return index
    return 0


def _is_numeric(value: object) -> bool:
    if isinstance(value, bool):
        return False
    try:
        float(_text(value).replace(" ", "").replace(",", ".").rstrip("%"))
        return bool(_text(value))
    except ValueError:
        return False


def _column_evidence(sheet: RawSheet, header: int, names: list[str]) -> list[dict[str, object]]:
    result = []
    rows = sheet.grid[header + 1 :]
    for column in range(sheet.n_cols):
        present = [row[column] for row in rows if _text(row[column])]
        unique = Counter(_text(value) for value in present)
        item: dict[str, object] = {
            "column_id": get_column_letter(column + 1),
            "header": names[column],
            "non_null": len(present),
            "numeric_count": sum(_is_numeric(value) for value in present),
            "samples": [_cap(value, defaults.EVIDENCE_CELL_CAP) for value in present[:3]],
        }
        if 0 < len(unique) <= defaults.EVIDENCE_UNIQUES_MAX:
            item["values"] = dict(unique.most_common())
        result.append(item)
    return result


def _dialogue_hints(columns: list[dict[str, object]]) -> list[dict[str, object]]:
    hints = []
    for column in columns:
        samples = [str(value) for value in column["samples"]]
        markers = Counter(marker for value in samples for marker in _BLOB_MARKER.findall(value))
        if len(markers) >= 2:
            hints.append({"column_id": column["column_id"], "markers": list(markers)})
    return hints


def build_sheet_evidence(sheet: RawSheet) -> dict[str, object]:
    header = _header_candidate(sheet)
    merge_map = header_merge_map(sheet, [header])
    names = [
        merge_header(sheet, [header], column, merge_map)
        or f"unnamed_{get_column_letter(column + 1)}"
        for column in range(sheet.n_cols)
    ]
    columns = _column_evidence(sheet, header, names)
    return {
        "sheet_name": sheet.name,
        "n_rows": sheet.n_rows,
        "n_cols": sheet.n_cols,
        "header_row_candidate": header + 1,
        "header_preview": [
            {"row": row + 1, "cells": [_cap(value, 80) for value in sheet.grid[row]]}
            for row in range(min(defaults.EVIDENCE_HEADER_PREVIEW_ROWS, sheet.n_rows))
        ],
        "columns": columns,
        "merged_ranges": [list(item) for item in sheet.merged],
        "formula_cells": [
            {"row": row + 1, "column_id": get_column_letter(column + 1), "formula": formula[:120]}
            for (row, column), formula in sorted(sheet.formulas.items())
        ],
        "dialogue_hints": _dialogue_hints(columns),
    }
