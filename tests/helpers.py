"""Помощники поведенческих тестов: сборка пакетов корзин и FakeClient."""
from __future__ import annotations

import json
import zipfile
from pathlib import Path

import openpyxl

_DOCX_SHELL = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
    "<w:body>{body}</w:body></w:document>"
)


def make_workbook(path: Path, sheets: dict) -> None:
    """sheets: {имя: {"rows": [[...]], "merges": ["A2:A3", ...]}}."""
    book = openpyxl.Workbook()
    book.remove(book.active)
    for name, spec in sheets.items():
        sheet = book.create_sheet(name)
        for row in spec["rows"]:
            sheet.append(row)
        for merge in spec.get("merges", []):
            sheet.merge_cells(merge)
    book.save(path)


def make_docx(path: Path, paragraphs=(), heading: str | None = None, table=None) -> None:
    parts = []
    if heading:
        parts.append(
            '<w:p><w:pPr><w:pStyle w:val="Heading1"/></w:pPr>'
            f"<w:r><w:t>{heading}</w:t></w:r></w:p>"
        )
    for text in paragraphs:
        parts.append(f"<w:p><w:r><w:t>{text}</w:t></w:r></w:p>")
    if table:
        rows = "".join(
            "<w:tr>"
            + "".join(f"<w:tc><w:p><w:r><w:t>{cell}</w:t></w:r></w:p></w:tc>" for cell in row)
            + "</w:tr>"
            for row in table
        )
        parts.append(f"<w:tbl>{rows}</w:tbl>")
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("word/document.xml", _DOCX_SHELL.format(body="".join(parts)))


def make_package(tmp_path: Path, sheets: dict, validation=(), development=(),
                 instruction: str = "Оценка: 1 — верно, 0 — неверно",
                 name: str = "CI09000001_test") -> Path:
    package = tmp_path / name
    package.mkdir()
    make_workbook(package / "test_set.xlsx", sheets)
    make_docx(package / "validation_report.docx",
              paragraphs=validation or ("Ключевая метрика Accuracy равна 0.5",),
              heading="Отчет о валидации")
    make_docx(package / "development_report.docx",
              paragraphs=development or ("Расшифровка колонок: q - запрос, a - ответ",),
              heading="Отчет о разработке")
    (package / "assessor_instruction.txt").write_text(instruction, encoding="utf-8")
    return package


class FakeClient:
    """Duck-type LlmClient: отдаёт заготовленные ответы по очереди, пишет вызовы."""

    def __init__(self, responses):
        self.responses = list(responses)
        self.labels: list[str] = []
        self.histories: list[list[dict]] = []
        self.structured_output = None
        self.calls = 0
        self.repair_turns = 0
        self.transport_retries = 0
        self.config = type("Cfg", (), {"model": "fake-model"})()

    def chat(self, messages, label, response_schema=None):
        self.labels.append(label)
        self.histories.append(list(messages))
        self.calls += 1
        item = self.responses.pop(0)
        return item if isinstance(item, str) else json.dumps(item, ensure_ascii=False)
