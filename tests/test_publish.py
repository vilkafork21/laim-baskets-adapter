"""Проекция внутреннего канона в формат тестового датасета."""

from __future__ import annotations

import ast
import re
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from laim_basket.errors import NotEvaluableError
from laim_basket.models import MeasurementPlan
from laim_basket.resolve import resolve_layout
from laim_basket.publish import DIALOGUE_SHEET, FLAT_SHEET, metric_slug, publish_umr
from laim_basket.reading.xlsx_reader import read_workbook
from laim_basket.transform.canon import build_canon
from laim_basket.transform.grouping import apply_grouping

from conftest import layout_proposal as _proposal
from conftest import source as _source


def _materialize(tmp_path: Path, header: list[str], rows: list[list[object]], proposal: dict):
    path = tmp_path / "basket.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet"
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    sheets = read_workbook(path)
    layout = resolve_layout(proposal, sheets, "CI1", "", frozenset())
    grouped = apply_grouping(sheets["Sheet"], layout.region, layout.transform_config())
    frame, _report = build_canon(grouped, layout.region, layout.transform_config())
    return layout, frame


def _plan(mode: str, sources: list[dict], reducer: str = "mean") -> MeasurementPlan:
    return MeasurementPlan(
        basket_id="CI1", metric_name="quality", assessment_mode=mode, method="identity", sources=tuple(sources),
        missing_policy="fail", majority_denominator=None, reducer=reducer,
        threshold=None, comparator=None, scale="ratio", precision=3,
        reported_value=None, reported_raw=None,
        evidence={"metric": ("doc-1:p0001",)},
    )


@pytest.mark.parametrize(
    ("header", "expected"),
    [
        ("mark1", "mark1_metric"),
        ("Полнота\nmarker1 ", "полнота_marker1_metric"),
        ("mark * freq", "mark_freq_metric"),
        ("Оценка классификации", "оценка_классификации_metric"),
        ("quality_metric", "quality_metric"),
    ],
)
def test_metric_slug_is_deterministic_from_header(header, expected):
    assert metric_slug(header) == expected


def test_flat_projection_publishes_only_spec_columns(tmp_path):
    header = ["Тема", "Вопрос", "Ответ", "Эталон", "Оценка", "Эксперт", "freq", "Комментарий"]
    rows = [
        ["кредит", "q1", "a1", "r1", "да", "ivanov", 2, "ok"],
        ["вклад", "q2", "a2", "r2", "нет", "petrov", 1, None],
    ]
    layout, frame = _materialize(tmp_path, header, rows, _proposal({
        "scenario": "A",
        "input_query": "B",
        "output_answer": "C",
        "reference_answers": ["D"],
        "assessor_id": "F",
    }, weight="G"))
    frame["main_metric"] = [1.0, 0.0]
    plan = _plan("qa", [_source("E", "final_score", {"да": 1, "нет": 0})])

    published = publish_umr(frame, layout, plan)

    assert published.variant == "flat"
    assert published.sheet_name == FLAT_SHEET
    assert list(published.frame.columns) == [
        "scenario", "query_id", "input_query_count", "input_query", "output_answer",
        "reference_answer", "оценка_metric", "main_metric", "assessor_id",
    ]
    assert published.frame["оценка_metric"].tolist() == [1.0, 0.0]
    assert published.frame["input_query_count"].tolist() == [2, 1]
    assert published.frame["query_id"].tolist() == ["row-0", "row-1"]
    assert published.published_columns == {
        "A": "scenario", "B": "input_query", "C": "output_answer", "D": "reference_answer",
        "E": "оценка_metric", "F": "assessor_id", "G": "input_query_count",
    }


def test_multiple_references_numbered_from_one(tmp_path):
    # Спецификация: несколько эталонов — reference_answer_1, reference_answer_2, …
    header = ["Вопрос", "Ответ", "Эталон", "Альтернатива", "Оценка"]
    rows = [["q1", "a1", "r1", "r1b", 1], ["q2", "a2", "r2", None, 0]]
    layout, frame = _materialize(tmp_path, header, rows, _proposal({
        "input_query": "A",
        "output_answer": "B",
        "reference_answers": ["C", "D"],
    }))
    plan = _plan("qa", [_source("E", "final_score")])
    frame["main_metric"] = [1.0, 0.0]

    published = publish_umr(frame, layout, plan)

    assert "reference_answer" not in published.frame.columns
    assert published.frame["reference_answer_1"].tolist() == ["r1", "r2"]
    assert published.frame["reference_answer_2"].tolist() == ["r1b", None]
    assert published.published_columns["C"] == "reference_answer_1"
    assert published.published_columns["D"] == "reference_answer_2"


def test_single_reference_keeps_bare_name(tmp_path):
    header = ["Вопрос", "Ответ", "Эталон", "Оценка"]
    rows = [["q1", "a1", "r1", 1]]
    layout, frame = _materialize(tmp_path, header, rows, _proposal({
        "input_query": "A",
        "output_answer": "B",
        "reference_answers": ["C"],
    }))
    plan = _plan("qa", [_source("D", "final_score")])
    frame["main_metric"] = [1.0]

    published = publish_umr(frame, layout, plan)

    assert published.published_columns["C"] == "reference_answer"


# Характеризационный контракт спеки «Формат тестового датасета»: публикуемые
# имена — только из её словаря. Тест фиксирует соответствие, а не меняет его.
_SPEC_COLUMN = re.compile(
    r"^(solution_version|scenario|session_id|query_id|input_query_count"
    r"|input_query|output_answer|assessor_id|main_metric|dialogue)$"
    r"|^reference_answer(_[0-9]+)?$"
    r"|^\w+_(input_query|output_answer|reference_answer)$"
    r"|^\w+_metric$"
)


def test_published_columns_stay_within_spec_vocabulary(tmp_path):
    header = ["Тема", "Вопрос", "Ответ", "Эталон", "Альтернатива", "Оценка",
              "Эксперт", "freq"]
    rows = [["кредит", "q1", "a1", "r1", "r1b", 1, "ivanov", 2],
            ["вклад", "q2", "a2", "r2", None, 0, "petrov", 1]]
    layout, frame = _materialize(tmp_path, header, rows, _proposal({
        "scenario": "A",
        "input_query": "B",
        "output_answer": "C",
        "reference_answers": ["D", "E"],
        "assessor_id": "G",
    }, weight="H"))
    plan = _plan("qa", [_source("F", "final_score")])
    frame["main_metric"] = [1.0, 0.0]

    published = publish_umr(frame, layout, plan)

    stray = [name for name in published.frame.columns if not _SPEC_COLUMN.match(str(name))]
    assert stray == []


def test_flat_projection_without_plan_has_no_metric_columns(tmp_path):
    layout, frame = _materialize(tmp_path, ["Q", "A", "S"], [["q1", "a1", 1]], _proposal({
        "input_query": "A", "output_answer": "B",
    }))

    published = publish_umr(frame, layout, None)

    assert list(published.frame.columns) == ["query_id", "input_query_count", "input_query", "output_answer"]
    assert published.published_columns == {"A": "input_query", "B": "output_answer"}


def test_turn_level_groups_publish_session_id_from_grouping_column(tmp_path):
    header = ["conversation_id", "Q", "A", "score"]
    rows = [["conv-1", "q1", "a1", 1], [None, "q2", "a2", 0], ["conv-2", "q3", "a3", 1]]
    layout, frame = _materialize(tmp_path, header, rows, _proposal(
        {"input_query": "B", "output_answer": "C"},
        grouping={"kind": "column", "column": "A"},
    ))
    frame["main_metric"] = [1.0, 0.0, 1.0]
    plan = _plan("turn_with_history", [_source("D", "final_score")])

    published = publish_umr(frame, layout, plan)

    assert published.variant == "flat"
    assert published.frame["session_id"].tolist() == ["conv-1", "conv-1", "conv-2"]
    assert published.frame["query_id"].tolist() == ["row-0", "row-1", "row-2"]
    assert published.published_columns["A"] == "session_id"
    assert "reference_group_id" not in published.frame and "turn_index" not in published.frame


def test_query_id_is_scoped_by_session_and_preserved(tmp_path):
    header = ["session", "query", "Q", "A", "score"]
    rows = [
        ["s1", "q1", "q1", "a1", 1],
        ["s2", "q1", "q2", "a2", 1],
    ]
    layout, frame = _materialize(tmp_path, header, rows, _proposal(
        {
            "session_id": "A",
            "query_id": "B",
            "input_query": "C",
            "output_answer": "D",
        },
        grouping={"kind": "column", "column": "A"},
    ))
    frame["main_metric"] = [1.0, 1.0]

    published = publish_umr(frame, layout, _plan("qa", [_source("E", "final_score")]))

    assert published.frame["session_id"].tolist() == ["s1", "s2"]
    assert published.frame["query_id"].tolist() == ["q1", "q1"]


def test_query_id_duplicate_inside_session_is_synthesized(tmp_path):
    proposal = _proposal(
        {
            "session_id": "A",
            "query_id": "B",
            "input_query": "C",
            "output_answer": "D",
        },
        grouping={"kind": "column", "column": "A"},
    )

    _layout, frame = _materialize(
        tmp_path,
        ["session", "query", "Q", "A"],
        [["s1", "q1", "q1", "a1"], ["s1", "q1", "q2", "a2"]],
        proposal,
    )

    assert frame["query_id"].tolist() == ["row-0", "row-1"]


def test_blob_dialogue_preserves_explicit_session_ids(tmp_path):
    session_ids = [
        "11111111-1111-4111-8111-111111111111",
        "22222222-2222-4222-8222-222222222222",
    ]
    rows = [
        [session_ids[0], repr([("q1", "q1", "a1")]), 1],
        [session_ids[1], repr([("q1", "q2", "a2")]), 1],
    ]
    layout, frame = _materialize(
        tmp_path,
        ["session", "dialogue", "mark"],
        rows,
        _proposal(
            {
                "session_id": "A",
                "input_query": "B",
                "output_answer": None,
            },
            grouping={"kind": "blob_row", "column": None},
            dialogue_blob={
                "column": "B",
                "container": "python_list",
                "question_marker": "КЛИЕНТ",
                "answer_marker": "АГЕНТ",
            },
        ),
    )
    frame["main_metric"] = [1.0, 1.0]

    published = publish_umr(
        frame,
        layout,
        _plan("dialogue", [_source("C", "final_score")]),
    )

    assert published.frame["session_id"].tolist() == session_ids
    assert [ast.literal_eval(value)[0][0] for value in published.frame["dialogue"]] == [
        "q1",
        "q1",
    ]


def test_dialogue_projection_packs_turns_and_keeps_session_columns(tmp_path):
    header = ["D", "Тема", "Q", "A", "mark1", "mark2", "Эксперт"]
    rows = [
        ["d1", "вклад", "q1", "a1", 1, 1, "ivanov"],
        [None, "вклад", "q2", "a2", 1, 1, "ivanov"],
        ["d2", "кредит", "q3", "a3", 0, 1, "petrov"],
    ]
    layout, frame = _materialize(tmp_path, header, rows, _proposal(
        {
            "scenario": "B", "input_query": "C",
            "output_answer": "D", "assessor_id": "G",
        },
        grouping={"kind": "column", "column": "A"},
    ))
    frame["main_metric"] = [1.0, 1.0, 0.0]
    plan = _plan("dialogue", [_source("E", "assessor_vote"), _source("F", "assessor_vote")])

    published = publish_umr(frame, layout, plan)

    assert published.variant == "dialogue"
    assert published.sheet_name == DIALOGUE_SHEET
    assert list(published.frame.columns) == [
        "scenario", "session_id", "dialogue", "mark1_metric", "mark2_metric",
        "main_metric", "assessor_id",
    ]
    assert published.frame["session_id"].tolist() == ["d1", "d2"]
    assert published.frame["scenario"].tolist() == ["вклад", "кредит"]
    assert ast.literal_eval(published.frame["dialogue"].iloc[0]) == [
        ("row-0", "q1", "a1"), ("row-1", "q2", "a2"),
    ]
    assert published.frame["mark1_metric"].tolist() == [1.0, 0.0]
    assert published.frame["main_metric"].tolist() == [1.0, 0.0]
    assert published.frame["assessor_id"].tolist() == ["ivanov", "petrov"]
    assert published.dropped_non_constant == ()


def test_weighted_dialogue_publishes_weight_column(tmp_path):
    header = ["D", "Q", "A", "score", "freq"]
    rows = [["d1", "q1", "a1", 1, 3], [None, "q2", "a2", None, 3], ["d2", "q3", "a3", 0, 1]]
    layout, frame = _materialize(tmp_path, header, rows, _proposal(
        {"input_query": "B", "output_answer": "C"},
        grouping={"kind": "column", "column": "A"}, weight="E",
    ))
    frame["main_metric"] = [1.0, 1.0, 0.0]
    plan = _plan("dialogue", [_source("D", "final_score")], reducer="frequency_weighted_mean")

    published = publish_umr(frame, layout, plan)

    assert list(published.frame.columns) == [
        "session_id", "dialogue", "input_query_count", "score_metric", "main_metric",
    ]
    assert published.frame["input_query_count"].tolist() == [3, 1]


def test_unweighted_dialogue_omits_weight_column(tmp_path):
    header = ["D", "Q", "A", "score", "freq"]
    rows = [["d1", "q1", "a1", 1, 3], ["d2", "q3", "a3", 0, 1]]
    layout, frame = _materialize(tmp_path, header, rows, _proposal(
        {"input_query": "B", "output_answer": "C"},
        grouping={"kind": "column", "column": "A"}, weight="E",
    ))
    frame["main_metric"] = [1.0, 0.0]

    published = publish_umr(frame, layout, _plan("dialogue", [_source("D", "final_score")]))

    assert "input_query_count" not in published.frame


def test_dialogue_projection_drops_session_columns_that_vary_inside_dialogue(tmp_path):
    header = ["D", "Тема", "Q", "A", "score"]
    rows = [["d1", "вклад", "q1", "a1", 1], ["d1", "кредит", "q2", "a2", 1]]
    layout, frame = _materialize(tmp_path, header, rows, _proposal(
        {"scenario": "B", "input_query": "C", "output_answer": "D"},
        grouping={"kind": "column", "column": "A"},
    ))
    frame["main_metric"] = [1.0, 1.0]

    published = publish_umr(frame, layout, _plan("dialogue", [_source("E", "final_score")]))

    assert "scenario" not in published.frame
    assert published.dropped_non_constant == ("scenario",)


def test_dialogue_with_merged_rows_synthesizes_integer_session_id(tmp_path):
    header = ["Q", "A", "score"]
    rows = [["q1", "a1", 1], ["q2", None, None], ["q3", "a3", 0]]
    path = tmp_path / "basket.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet"
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    sheet.merge_cells("B2:B3")
    sheet.merge_cells("C2:C3")
    workbook.save(path)
    sheets = read_workbook(path)
    layout = resolve_layout(_proposal(
        {"input_query": "A", "output_answer": "B"},
        grouping={"kind": "merged_rows", "column": None},
    ), sheets, "CI1", "", frozenset())
    grouped = apply_grouping(sheets["Sheet"], layout.region, layout.transform_config())
    frame, _ = build_canon(grouped, layout.region, layout.transform_config())
    frame["main_metric"] = [1.0, 1.0, 0.0]

    published = publish_umr(frame, layout, _plan("dialogue", [_source("C", "final_score")]))

    assert published.frame["session_id"].tolist() == [1, 2]
    assert ast.literal_eval(published.frame["dialogue"].iloc[0]) == [("row-0", "q1", "a1"), ("row-1", "q2", "a1")]


def test_accuracy_sources_get_own_module_columns(tmp_path):
    header = ["request", "class", "GT"]
    layout, frame = _materialize(tmp_path, header, [["q1", "a", "a"]], _proposal({
        "input_query": "A", "output_answer": "B",
        "reference_answers": ["C"],
    }))
    frame["main_metric"] = [1.0]
    plan = _plan("qa", [_source("B", "prediction", "label"), _source("C", "target", "label")])

    published = publish_umr(frame, layout, plan)

    # Монитор поставляет предсказание отдельным полем, поэтому источники accuracy
    # получают собственные имена даже на колонках ответа и эталона.
    assert published.published_columns == {
        "A": "input_query", "B": "class_output_answer", "C": "gt_reference_answer",
    }
    assert list(published.frame.columns) == [
        "query_id", "input_query_count", "input_query", "output_answer",
        "class_output_answer", "reference_answer", "gt_reference_answer", "main_metric",
    ]
    assert published.frame["class_output_answer"].tolist() == ["a"]
    assert published.frame["gt_reference_answer"].tolist() == ["a"]


def test_dialogue_publishes_label_sources_named_in_the_contract(tmp_path):
    header = ["D", "Q", "A", "route", "gold"]
    rows = [["d1", "q1", "a1", "picker", "picker"], ["d1", "q2", "a2", "picker", "picker"]]
    layout, frame = _materialize(tmp_path, header, rows, _proposal(
        {"input_query": "B", "output_answer": "C"},
        grouping={"kind": "column", "column": "A"},
    ))
    frame["main_metric"] = [1.0, 1.0]
    plan = _plan("dialogue", [_source("D", "prediction", "label"), _source("E", "target", "label")])

    published = publish_umr(frame, layout, plan)

    assert published.variant == "dialogue"
    assert list(published.frame.columns) == [
        "session_id", "dialogue", "route_output_answer", "gold_reference_answer",
        "main_metric",
    ]
    assert published.frame["route_output_answer"].tolist() == ["picker"]


@pytest.mark.parametrize(
    ("mode", "expected_columns", "expected_versions"),
    [
        (
            "qa",
            [
                "solution_version", "session_id", "query_id", "input_query_count",
                "input_query", "output_answer", "score_metric", "main_metric",
            ],
            ["D-01.002.03", "D-01.002.03"],
        ),
        (
            "dialogue",
            ["solution_version", "session_id", "dialogue", "score_metric", "main_metric"],
            ["D-01.002.03"],
        ),
    ],
)
def test_solution_version_is_preserved_in_each_output_format(
    tmp_path, mode, expected_columns, expected_versions
):
    header = ["solution_version", "D", "Q", "A", "score"]
    rows = [
        ["D-01.002.03", "d1", "q1", "a1", 1],
        ["D-01.002.03", "d1", "q2", "a2", 1],
    ]
    layout, frame = _materialize(tmp_path, header, rows, _proposal(
        {"input_query": "C", "output_answer": "D"},
        grouping={"kind": "column", "column": "B"},
    ))
    frame["main_metric"] = [1.0, 1.0]

    published = publish_umr(frame, layout, _plan(mode, [_source("E", "final_score")]))

    assert list(published.frame.columns) == expected_columns
    assert published.frame["solution_version"].tolist() == expected_versions
    assert published.published_columns["A"] == "solution_version"


def test_source_on_a_role_column_is_rejected(tmp_path):
    layout, frame = _materialize(tmp_path, ["Q", "A", "Тема"], [["q1", "a1", "x"]], _proposal({
        "input_query": "A", "output_answer": "B",
        "scenario": "C",
    }))
    frame["main_metric"] = [1.0]
    with pytest.raises(NotEvaluableError) as excinfo:
        publish_umr(frame, layout, _plan("qa", [_source("C", "final_score")]))
    assert excinfo.value.details["column_id"] == "C"


def test_colliding_metric_names_are_rejected(tmp_path):
    layout, frame = _materialize(tmp_path, ["Q", "A", "mark 1", "mark_1"], [["q1", "a1", 1, 1]], _proposal({
        "input_query": "A", "output_answer": "B",
    }))
    frame["main_metric"] = [1.0]
    with pytest.raises(NotEvaluableError):
        publish_umr(frame, layout, _plan("qa", [_source("C", "criterion"), _source("D", "criterion")]))


def test_value_map_criterion_is_published_as_number(tmp_path):
    layout, frame = _materialize(tmp_path, ["Q", "A", "verdict"], [["q1", "a1", "да"], ["q2", "a2", "нет"]], _proposal({
        "input_query": "A", "output_answer": "B",
    }))
    frame["main_metric"] = [1.0, 0.0]

    published = publish_umr(frame, layout, _plan("qa", [_source("C", "final_score", {"да": 1, "нет": 0})]))

    assert published.frame["verdict_metric"].tolist() == [1.0, 0.0]


def test_published_metric_values_are_normalized_numbers(tmp_path):
    layout, frame = _materialize(tmp_path, ["Q", "A", "score"], [["q1", "a1", "50%"], ["q2", "a2", "1,5"]], _proposal({
        "input_query": "A", "output_answer": "B",
    }))
    frame["main_metric"] = [0.5, 1.5]
    published = publish_umr(frame, layout, _plan("qa", [_source("C", "final_score")]))
    assert published.frame["score_metric"].tolist() == [0.5, 1.5]
    assert all(isinstance(value, float) for value in published.frame["score_metric"])
    assert Decimal("0.5") == Decimal(str(published.frame["score_metric"].iloc[0]))
